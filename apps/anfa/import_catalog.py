"""Parse the clinic's Excel service export and load it into the catalog.

The clinic exports its full service list from their local CRM as an `.xlsx`
with a single `services` sheet and columns: tab, category_name, title, price.
They re-export on every change and hand us the new file; we treat each import
as authoritative and reconcile the whole catalog against it
(`AnfaRepository.replace_catalog` upserts the current rows and drops any that
disappeared).

Import-light: `openpyxl` is imported lazily inside `parse_workbook` so importing
this module (e.g. from the admin panel) stays cheap, mirroring the rest of the
`apps/anfa` data layer.
"""

from __future__ import annotations

import io
import logging
from typing import Optional

logger = logging.getLogger(__name__)

# Header names we accept, mapped to our canonical field. The export uses the
# first form; the aliases tolerate light renames without breaking the import.
_COLUMN_ALIASES = {
    "tab": "tab",
    "category_name": "category",
    "category": "category",
    "title": "title",
    "name": "title",
    "price": "price",
}

_SHEET_NAME = "services"


def _coerce_price(raw) -> int:
    """Best-effort price → whole UZS sum. Blanks / non-numeric → 0."""
    if raw is None:
        return 0
    if isinstance(raw, (int, float)):
        return int(raw)
    digits = "".join(ch for ch in str(raw) if ch.isdigit())
    return int(digits) if digits else 0


def parse_workbook(data: bytes) -> list[dict]:
    """Parse xlsx bytes into a list of {tab, category, title, price} dicts.

    Reads the `services` sheet (falls back to the first sheet), maps columns by
    header name, skips rows without a title, and de-duplicates identical
    (tab, category, title) rows keeping the last price seen. Raises ValueError
    on a file that has no recognisable title column.
    """
    import openpyxl  # lazy — keeps module import cheap

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.worksheets[0]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []

    # Map column index -> canonical field via the header labels.
    col_field: dict[int, str] = {}
    for idx, label in enumerate(header):
        key = str(label).strip().lower() if label is not None else ""
        if key in _COLUMN_ALIASES:
            col_field[idx] = _COLUMN_ALIASES[key]
    if "title" not in col_field.values():
        raise ValueError(
            "Excel export has no 'title' column — expected headers "
            "tab, category_name, title, price."
        )

    by_key: dict[tuple, dict] = {}
    for row in rows_iter:
        rec = {"tab": "", "category": "", "title": "", "price": 0}
        for idx, field in col_field.items():
            value = row[idx] if idx < len(row) else None
            if field == "price":
                rec["price"] = _coerce_price(value)
            else:
                rec[field] = (str(value).strip() if value is not None else "")
        if not rec["title"]:
            continue
        by_key[(rec["tab"], rec["category"], rec["title"])] = rec

    return list(by_key.values())


async def import_workbook_bytes(data: bytes, repo: Optional[object] = None) -> dict:
    """Parse an export and reconcile it into the catalog. Returns the
    `replace_catalog` summary plus the parsed row count."""
    from apps.anfa.repository import get_repository

    rows = parse_workbook(data)
    repository = repo or get_repository()
    summary = await repository.replace_catalog(rows)
    summary["parsed"] = len(rows)
    logger.info(
        "Catalog import: parsed=%d added=%d updated=%d removed=%d total=%d",
        summary["parsed"],
        summary["added"],
        summary["updated"],
        summary["removed"],
        summary["total"],
    )
    return summary
